from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from models import Peminjaman, Pengembalian
from app_instance import db
from datetime import datetime, timedelta
import io

riwayat_bp = Blueprint('riwayat', __name__)


def parse_date_filter(filter_type, date_from=None, date_to=None):
    today = datetime.now().date()
    if filter_type == 'today':
        return today, today
    elif filter_type == '7days':
        return today - timedelta(days=7), today
    elif filter_type == '30days':
        return today - timedelta(days=30), today
    elif filter_type == 'custom' and date_from and date_to:
        return datetime.strptime(date_from, '%Y-%m-%d').date(), datetime.strptime(date_to, '%Y-%m-%d').date()
    return None, None


@riwayat_bp.route('', methods=['GET'])
@jwt_required()
def get_riwayat():
    filter_type = request.args.get('filter', 'today')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    record_type = request.args.get('type', 'all')  # all, peminjaman, pengembalian

    start, end = parse_date_filter(filter_type, date_from, date_to)

    result = []

    if record_type in ('all', 'peminjaman'):
        q = Peminjaman.query
        if start and end:
            q = q.filter(db.func.date(Peminjaman.borrowed_at) >= start,
                         db.func.date(Peminjaman.borrowed_at) <= end)
        for p in q.order_by(Peminjaman.borrowed_at.desc()).all():
            d = p.to_dict()
            d['record_type'] = 'Peminjaman'
            d['waktu'] = d['borrowed_at']
            result.append(d)

    if record_type in ('all', 'pengembalian'):
        q = Pengembalian.query
        if start and end:
            q = q.filter(db.func.date(Pengembalian.returned_at) >= start,
                         db.func.date(Pengembalian.returned_at) <= end)
        for r in q.order_by(Pengembalian.returned_at.desc()).all():
            d = r.to_dict()
            d['record_type'] = 'Pengembalian'
            d['waktu'] = d['returned_at']
            result.append(d)

    result.sort(key=lambda x: x.get('waktu') or '', reverse=True)
    return jsonify(result), 200


@riwayat_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filter_type = request.args.get('filter', 'today')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    record_type = request.args.get('type', 'all')
    start, end = parse_date_filter(filter_type, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Inventori"

    # Header styling
    header_fill = PatternFill(start_color="7B5EA7", end_color="7B5EA7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['No', 'Tipe', 'Nama', 'Divisi', 'Barang', 'Catatan Kondisi', 'Waktu']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Collect data
    rows = []
    if record_type in ('all', 'peminjaman'):
        q = Peminjaman.query
        if start and end:
            q = q.filter(db.func.date(Peminjaman.borrowed_at) >= start,
                         db.func.date(Peminjaman.borrowed_at) <= end)
        for p in q.all():
            rows.append([
                'Peminjaman', p.borrower_name, p.division,
                p.item.name if p.item else '-',
                p.condition_note or '-',
                p.borrowed_at.strftime('%d/%m/%Y %H:%M') if p.borrowed_at else '-'
            ])

    if record_type in ('all', 'pengembalian'):
        q = Pengembalian.query
        if start and end:
            q = q.filter(db.func.date(Pengembalian.returned_at) >= start,
                         db.func.date(Pengembalian.returned_at) <= end)
        for r in q.all():
            rows.append([
                'Pengembalian', r.returner_name, r.division,
                r.peminjaman.item.name if r.peminjaman and r.peminjaman.item else '-',
                r.condition_note or '-',
                r.returned_at.strftime('%d/%m/%Y %H:%M') if r.returned_at else '-'
            ])

    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 2):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = border
        ws.cell(row=i, column=1, value=i - 1).border = border

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='riwayat_inventori.xlsx')


@riwayat_bp.route('/export/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    filter_type = request.args.get('filter', 'today')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    record_type = request.args.get('type', 'all')
    start, end = parse_date_filter(filter_type, date_from, date_to)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Riwayat Inventori R&D & Teknisi", styles['Title']))
    elements.append(Paragraph(f"PT Interskala Mandiri Indonesia", styles['Normal']))
    elements.append(Spacer(1, 12))

    rows = [['No', 'Tipe', 'Nama', 'Divisi', 'Barang', 'Catatan', 'Waktu']]

    if record_type in ('all', 'peminjaman'):
        q = Peminjaman.query
        if start and end:
            q = q.filter(db.func.date(Peminjaman.borrowed_at) >= start,
                         db.func.date(Peminjaman.borrowed_at) <= end)
        for i, p in enumerate(q.all(), 1):
            rows.append([
                str(i), 'Peminjaman', p.borrower_name, p.division,
                p.item.name if p.item else '-',
                (p.condition_note or '-')[:30],
                p.borrowed_at.strftime('%d/%m/%Y %H:%M') if p.borrowed_at else '-'
            ])

    if record_type in ('all', 'pengembalian'):
        q = Pengembalian.query
        if start and end:
            q = q.filter(db.func.date(Pengembalian.returned_at) >= start,
                         db.func.date(Pengembalian.returned_at) <= end)
        for i, r in enumerate(q.all(), 1):
            rows.append([
                str(i), 'Pengembalian', r.returner_name, r.division,
                r.peminjaman.item.name if r.peminjaman and r.peminjaman.item else '-',
                (r.condition_note or '-')[:30],
                r.returned_at.strftime('%d/%m/%Y %H:%M') if r.returned_at else '-'
            ])

    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7B5EA7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F0F9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name='riwayat_inventori.pdf')
