from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from models import LaporanHarian
from app_instance import db
from datetime import datetime, timedelta
import io

laporan_bp = Blueprint('laporan', __name__)

HARI = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']


def _parse_time(value):
    if not value:
        return None
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    return None


def parse_date_filter(filter_type, date_from=None, date_to=None):
    today = datetime.now().date()
    if filter_type == 'today':
        return today, today
    elif filter_type == '7days':
        return today - timedelta(days=7), today
    elif filter_type == '30days':
        return today - timedelta(days=30), today
    elif filter_type == 'custom' and date_from and date_to:
        return datetime.strptime(date_from, '%Y-%m-%d').date(), datetime.strptime(date_to, '%Y-%m-%d').date()
    return None, None


@laporan_bp.route('', methods=['GET'])
@jwt_required()
def get_laporan():
    filter_type = request.args.get('filter', 'today')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    start, end = parse_date_filter(filter_type, date_from, date_to)
    claims = get_jwt()
    user_id = int(get_jwt_identity())

    q = LaporanHarian.query
    if claims.get('role') != 'admin':
        q = q.filter(LaporanHarian.teknisi_id == user_id)
    if start and end:
        q = q.filter(LaporanHarian.tanggal >= start, LaporanHarian.tanggal <= end)

    records = q.order_by(LaporanHarian.tanggal.desc()).all()
    return jsonify([r.to_dict() for r in records]), 200


@laporan_bp.route('', methods=['POST'])
@jwt_required()
def create_laporan():
    user_id = int(get_jwt_identity())
    data = request.get_json()
    tanggal = datetime.strptime(data['tanggal'], '%Y-%m-%d').date()
    hari = HARI[tanggal.weekday()]

    lap = LaporanHarian(
        tanggal=tanggal,
        hari=hari,
        pekerjaan=data['pekerjaan'],
        produk=data.get('produk'),
        informasi_tambahan=data.get('informasi_tambahan'),
        status_garansi=data.get('status_garansi', False),
        jam_start=_parse_time(data.get('jam_start')),
        jam_end=_parse_time(data.get('jam_end')),
        foto_url=data.get('foto_url'),
        teknisi_id=user_id
    )
    db.session.add(lap)
    db.session.commit()
    return jsonify(lap.to_dict()), 201


@laporan_bp.route('/<int:lid>', methods=['PUT'])
@jwt_required()
def update_laporan(lid):
    lap = LaporanHarian.query.get_or_404(lid)
    claims = get_jwt()
    user_id = int(get_jwt_identity())
    if claims.get('role') != 'admin' and lap.teknisi_id != user_id:
        return jsonify({'message': 'Anda tidak berhak mengubah laporan ini'}), 403
    data = request.get_json()
    if data.get('tanggal'):
        tanggal = datetime.strptime(data['tanggal'], '%Y-%m-%d').date()
        lap.tanggal = tanggal
        lap.hari = HARI[tanggal.weekday()]
    for f in ['pekerjaan', 'produk', 'informasi_tambahan', 'status_garansi', 'foto_url']:
        if f in data:
            setattr(lap, f, data[f])
    if data.get('jam_start'):
        lap.jam_start = _parse_time(data['jam_start'])
    if data.get('jam_end'):
        lap.jam_end = _parse_time(data['jam_end'])
    db.session.commit()
    return jsonify(lap.to_dict()), 200


@laporan_bp.route('/<int:lid>', methods=['DELETE'])
@jwt_required()
def delete_laporan(lid):
    lap = LaporanHarian.query.get_or_404(lid)
    claims = get_jwt()
    user_id = int(get_jwt_identity())
    if claims.get('role') != 'admin' and lap.teknisi_id != user_id:
        return jsonify({'message': 'Anda tidak berhak menghapus laporan ini'}), 403
    db.session.delete(lap)
    db.session.commit()
    return jsonify({'message': 'Laporan dihapus'}), 200


@laporan_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filter_type = request.args.get('filter', 'today')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    start, end = parse_date_filter(filter_type, date_from, date_to)
    claims = get_jwt()
    user_id = int(get_jwt_identity())

    q = LaporanHarian.query
    if claims.get('role') != 'admin':
        q = q.filter(LaporanHarian.teknisi_id == user_id)
    if start and end:
        q = q.filter(LaporanHarian.tanggal >= start, LaporanHarian.tanggal <= end)
    records = q.order_by(LaporanHarian.tanggal.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Harian Teknisi"

    header_fill = PatternFill(start_color="7B5EA7", end_color="7B5EA7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['No', 'Tanggal', 'Hari', 'Teknisi', 'Pekerjaan', 'Produk',
               'Info Tambahan', 'Garansi', 'Jam Mulai', 'Jam Selesai']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, r in enumerate(records, 2):
        row_data = [
            i - 1,
            r.tanggal.strftime('%d/%m/%Y') if r.tanggal else '-',
            r.hari or '-',
            r.teknisi.full_name if r.teknisi else '-',
            r.pekerjaan,
            r.produk or '-',
            r.informasi_tambahan or '-',
            'Ya' if r.status_garansi else 'Tidak',
            str(r.jam_start) if r.jam_start else '-',
            str(r.jam_end) if r.jam_end else '-'
        ]
        for j, val in enumerate(row_data, 1):
            ws.cell(row=i, column=j, value=val).border = border

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='laporan_harian.xlsx')


@laporan_bp.route('/export/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    filter_type = request.args.get('filter', 'today')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    start, end = parse_date_filter(filter_type, date_from, date_to)
    claims = get_jwt()
    user_id = int(get_jwt_identity())

    q = LaporanHarian.query
    if claims.get('role') != 'admin':
        q = q.filter(LaporanHarian.teknisi_id == user_id)
    if start and end:
        q = q.filter(LaporanHarian.tanggal >= start, LaporanHarian.tanggal <= end)
    records = q.order_by(LaporanHarian.tanggal.desc()).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Laporan Harian Teknisi", styles['Title']),
        Paragraph("PT Interskala Mandiri Indonesia", styles['Normal']),
        Spacer(1, 12)
    ]

    rows = [['No', 'Tanggal', 'Hari', 'Teknisi', 'Pekerjaan', 'Produk', 'Garansi', 'Jam']]
    for i, r in enumerate(records, 1):
        rows.append([
            str(i),
            r.tanggal.strftime('%d/%m/%Y') if r.tanggal else '-',
            r.hari or '-',
            r.teknisi.full_name if r.teknisi else '-',
            r.pekerjaan[:40] if r.pekerjaan else '-',
            r.produk or '-',
            'Ya' if r.status_garansi else 'Tidak',
            f"{r.jam_start or '-'} - {r.jam_end or '-'}"
        ])

    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7B5EA7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F0F9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name='laporan_harian.pdf')