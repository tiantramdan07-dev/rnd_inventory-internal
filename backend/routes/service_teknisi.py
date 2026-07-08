from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from functools import wraps
from models import ServiceTeknisi
from app_instance import db
from datetime import datetime, timedelta
import io

service_bp = Blueprint('service', __name__)


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        claims = get_jwt()
        if claims.get('role') != 'admin':
            return jsonify({'message': 'Hanya Admin yang dapat melakukan aksi ini'}), 403
        return fn(*args, **kwargs)
    return wrapper


def _generate_nomor():
    today = datetime.now()
    prefix = f"SRV/{today.strftime('%Y%m')}"
    count = ServiceTeknisi.query.filter(
        ServiceTeknisi.nomor.like(f"{prefix}%")
    ).count() + 1
    return f"{prefix}/{count:04d}"


def parse_date_filter(filter_type, date_from=None, date_to=None):
    today = datetime.now().date()
    if filter_type == 'today':
        return today, today
    elif filter_type == '7days':
        return today - timedelta(days=7), today
    elif filter_type == '30days':
        return today - timedelta(days=30), today
    elif filter_type == 'custom' and date_from and date_to:
        return datetime.strptime(date_from, '%Y-%m-%d').date(), datetime.strptime(date_to, '%Y-%m-%d').date()
    return None, None


@service_bp.route('', methods=['GET'])
@jwt_required()
def get_services():
    filter_type = request.args.get('filter', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    claims = get_jwt()
    user_id = int(get_jwt_identity())

    q = ServiceTeknisi.query
    if claims.get('role') != 'admin':
        q = q.filter(ServiceTeknisi.created_by == user_id)
    if filter_type != 'all':
        start, end = parse_date_filter(filter_type, date_from, date_to)
        if start and end:
            q = q.filter(ServiceTeknisi.tanggal_pengajuan >= start,
                         ServiceTeknisi.tanggal_pengajuan <= end)

    services = q.order_by(ServiceTeknisi.tanggal_pengajuan.desc()).all()
    return jsonify([s.to_dict() for s in services]), 200


@service_bp.route('', methods=['POST'])
@jwt_required()
def create_service():
    user_id = int(get_jwt_identity())
    from models import User
    current_user = User.query.get(user_id)
    data = request.get_json()
    s = ServiceTeknisi(
        nomor=_generate_nomor(),
        tanggal_pengajuan=datetime.strptime(data['tanggal_pengajuan'], '%Y-%m-%d').date(),
        nama_teknisi=current_user.full_name if current_user else data.get('nama_teknisi'),
        nama_sales=data.get('nama_sales'),
        customer=data['customer'],
        produk=data['produk'],
        lokasi=data.get('lokasi'),
        tipe_service=data.get('tipe_service'),
        tanggal_service=datetime.strptime(data['tanggal_service'], '%Y-%m-%d').date() if data.get('tanggal_service') else None,
        tanggal_selesai=datetime.strptime(data['tanggal_selesai'], '%Y-%m-%d').date() if data.get('tanggal_selesai') else None,
        status=data.get('status', 'Pending'),
        keterangan=data.get('keterangan'),
        created_by=user_id
    )
    db.session.add(s)
    db.session.commit()
    return jsonify(s.to_dict()), 201


@service_bp.route('/<int:sid>', methods=['PUT'])
@jwt_required()
@admin_required
def update_service(sid):
    s = ServiceTeknisi.query.get_or_404(sid)
    data = request.get_json()
    for field in ['nama_teknisi', 'nama_sales', 'customer', 'produk', 'lokasi', 'tipe_service', 'status', 'keterangan']:
        if field in data:
            setattr(s, field, data[field])
    if data.get('tanggal_pengajuan'):
        s.tanggal_pengajuan = datetime.strptime(data['tanggal_pengajuan'], '%Y-%m-%d').date()
    if data.get('tanggal_service'):
        s.tanggal_service = datetime.strptime(data['tanggal_service'], '%Y-%m-%d').date()
    if data.get('tanggal_selesai'):
        s.tanggal_selesai = datetime.strptime(data['tanggal_selesai'], '%Y-%m-%d').date()
    db.session.commit()
    return jsonify(s.to_dict()), 200


@service_bp.route('/<int:sid>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_service(sid):
    s = ServiceTeknisi.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    return jsonify({'message': 'Data service dihapus'}), 200


@service_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filter_type = request.args.get('filter', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    claims = get_jwt()
    user_id = int(get_jwt_identity())

    q = ServiceTeknisi.query
    if claims.get('role') != 'admin':
        q = q.filter(ServiceTeknisi.created_by == user_id)
    if filter_type != 'all':
        start, end = parse_date_filter(filter_type, date_from, date_to)
        if start and end:
            q = q.filter(ServiceTeknisi.tanggal_pengajuan >= start,
                         ServiceTeknisi.tanggal_pengajuan <= end)

    services = q.order_by(ServiceTeknisi.tanggal_pengajuan.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Service Teknisi"

    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['No', 'Nomor', 'Tgl Pengajuan', 'Teknisi', 'Sales', 'Customer',
               'Produk', 'Lokasi', 'Tipe', 'Tgl Service', 'Tgl Selesai', 'Status', 'Keterangan']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, s in enumerate(services, 2):
        row = [i - 1, s.nomor,
               s.tanggal_pengajuan.strftime('%d/%m/%Y') if s.tanggal_pengajuan else '-',
               s.nama_teknisi, s.nama_sales or '-', s.customer, s.produk, s.lokasi or '-',
               s.tipe_service or '-',
               s.tanggal_service.strftime('%d/%m/%Y') if s.tanggal_service else '-',
               s.tanggal_selesai.strftime('%d/%m/%Y') if s.tanggal_selesai else '-',
               s.status, s.keterangan or '-']
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val).border = border

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='service_teknisi.xlsx')


@service_bp.route('/export/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    filter_type = request.args.get('filter', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    claims = get_jwt()
    user_id = int(get_jwt_identity())

    q = ServiceTeknisi.query
    if claims.get('role') != 'admin':
        q = q.filter(ServiceTeknisi.created_by == user_id)
    if filter_type != 'all':
        start, end = parse_date_filter(filter_type, date_from, date_to)
        if start and end:
            q = q.filter(ServiceTeknisi.tanggal_pengajuan >= start,
                         ServiceTeknisi.tanggal_pengajuan <= end)

    services = q.order_by(ServiceTeknisi.tanggal_pengajuan.desc()).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Laporan Service Teknisi", styles['Title']),
        Paragraph("PT Interskala Mandiri Indonesia", styles['Normal']),
        Spacer(1, 12)
    ]

    rows = [['No', 'Nomor', 'Tgl', 'Teknisi', 'Customer', 'Produk', 'Tipe', 'Status']]
    for i, s in enumerate(services, 1):
        rows.append([
            str(i), s.nomor,
            s.tanggal_pengajuan.strftime('%d/%m/%Y') if s.tanggal_pengajuan else '-',
            s.nama_teknisi, s.customer, s.produk,
            s.tipe_service or '-', s.status
        ])

    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name='service_teknisi.pdf')